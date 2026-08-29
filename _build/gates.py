#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gates do padrão de treinamento.

Cada gate roda contra todas as páginas publicadas E é provado contra um defeito
injetado numa cópia em memória. Gate que não acusa o próprio defeito é gate que
não existe, e o script falha por isso, não só por achar problema no site.

🔴 Gate com exceção permanente deixa de ser gate. Se um gate acusa, conserta o
código, não o gate.

🔴 Exit code sozinho nunca é prova. Leia a saída: ela imprime achado por gate e
diz quais gates não se provaram contra o próprio defeito.

Rodar:  python3 _build/gates.py
"""

import io
import os
import re
import sys
import glob
import json
import importlib.util

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)

# A capa é o contrato do dia: lá a duração é informação do que foi combinado.
# Nas páginas de conteúdo ela vira cobrança para quem lê mais devagar.
PAGINAS_COM_AGENDA = {"index.html"}


def paginas():
    """Todo .html publicado, com caminho relativo à raiz do site."""
    achados = []
    for p in glob.glob(os.path.join(RAIZ, "**", "*.html"), recursive=True):
        rel = os.path.relpath(p, RAIZ)
        # _build é fonte, não página publicada. Auditar fonte é auditar arquivo
        # que ninguém abre, e o achado lá não aparece na tela de ninguém.
        if rel.split(os.sep)[0].startswith("_"):
            continue
        achados.append((rel, io.open(p, encoding="utf-8").read()))
    return sorted(achados)


def linhas_com(texto, padrao, flags=re.I):
    """(nº da linha, trecho) de cada linha inteira que casa.

    Nunca recorta antes de filtrar: recortar antes é como um gate de "pede nome"
    deu falso positivo em cima de "não escreva seu nome".
    """
    saida = []
    for n, linha in enumerate(texto.split("\n"), 1):
        if re.search(padrao, linha, flags):
            saida.append((n, linha.strip()[:150]))
    return saida


# Só tag de bloco. strong, em, span e a são inline: quebrar neles partiria a
# frase ao meio e faria todo gate de "na mesma frase" acusar sozinho.
BLOCOS = ("p", "div", "li", "h1", "h2", "h3", "h4", "td", "th", "tr",
          "section", "article", "summary", "label")


def texto_visivel(html):
    """Só o que o aluno lê: sem <style>, <script>, comentário e tag.

    Preserva a quebra entre blocos. Colapsar tudo numa linha faz o gate perder a
    noção de "na mesma frase" e devolver a página inteira como mensagem de erro.
    """
    s = re.sub(r"<style[^>]*>.*?</style>", " ", html, flags=re.S)
    s = re.sub(r"<script[^>]*>.*?</script>", " ", s, flags=re.S)
    s = re.sub(r"<!--.*?-->", " ", s, flags=re.S)
    # O espaço rígido que o gerador cola é espaço para quem lê. Se ele chegasse
    # aqui como caractere próprio, todo gate que compara frase passaria a acusar
    # a si mesmo depois da correção de quebra de linha.
    s = s.replace("&nbsp;", " ").replace(" ", " ")
    # Colapsa ANTES de marcar bloco: quebra de linha no arquivo-fonte não é
    # quebra na tela, e tratar as duas como iguais parte frase no meio.
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"</(" + "|".join(BLOCOS) + r")>", "\n", s)
    s = re.sub(r"<br\s*/?>", "\n", s)
    s = re.sub(r"<[^>]+>", " ", s)
    linhas = [re.sub(r"[ \t]+", " ", l).strip() for l in s.split("\n")]
    return "\n".join(l for l in linhas if l)


def e_vitrine(rel):
    """A vitrine é o único arquivo que ENUNCIA as regras em vez de obedecer a elas.

    Ela cita "pergunte à sala" para dizer que isso é proibido, e cita "40 minutos
    numa aula" para explicar por que duração não vai para a página do aluno. Gate
    de palavra rodando contra ela se auto-reprova.

    A exceção é estrutural, não caso a caso: a vitrine é referência interna e não
    vai para turma nenhuma. Se um dia ela for para a turma, esta linha cai junto.
    """
    return rel.split(os.sep)[0] == "componentes"


def blocos_por_classe(html, classe):
    """Todo <div class="... classe ..."> com o miolo dele, contando a profundidade
    das divs.

    Regex de vizinhança não serve: (.*?)</div> fecha na primeira div interna, e o
    gate passa a auditar o cabeçalho do componente em vez do componente. Foi assim
    que os dois gates das peças interativas nasceram cegos, com 0 achado e 0 prova.
    """
    achados = []
    for m in re.finditer(r'<div class="([^"]*)"([^>]*)>', html):
        if classe not in m.group(1).split():
            continue
        prof, i = 1, m.end()
        while prof and i < len(html):
            t = re.search(r"<(/?)div\b", html[i:])
            if not t:
                break
            i += t.end()
            prof += -1 if t.group(1) else 1
        achados.append((m.group(2), html[m.end():i - 6]))
    return achados


def css_da_pagina(html):
    return " ".join(re.findall(r"<style[^>]*>(.*?)</style>", html, flags=re.S))


# O gerador é importado, não copiado: gate que reimplementa a regra do gerador
# passa a testar a própria cópia, e as duas divergem sem ninguém ver.
_spec = importlib.util.spec_from_file_location(
    "gerador_do_treinamento", os.path.join(AQUI, "gerar.py"))
_gerador = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_gerador)

MARCA = io.open(os.path.join(AQUI, "marca.css"), encoding="utf-8").read()


# =========================================================================
# ESCRITA
# =========================================================================

def g1_travessao(rel, html):
    """Travessão é proibido em todo material do Rafael."""
    return ["linha com travessão: " + l
            for _, l in linhas_com(texto_visivel(html), r"—")]


VOCAB_INTERNO = [
    r"\bonda \d", r"\bgate\b", r"\bhandoff\b", r"\bspec\b", r"\bcommit\b",
    r"\bsubagente\b", r"\bclaude code\b", r"\bfrontmatter\b", r"\bfio[s]? em aberto\b",
    r"\bprompt de sistema\b", r"\bfragmento\b", r"\bbase\.css\b", r"\bgerar\.py\b",
]


def g2_vocabulario_interno(rel, html):
    """Vocabulário de bastidor não vai para a tela do aluno.

    A vitrine é a exceção declarada e única: ela existe justamente para nomear a
    peça e a regra dela, e é uso interno, não material de turma.
    """
    if e_vitrine(rel):
        return []
    # O curso declara em CURSO["ensina"] os nomes oficiais que ele ensina, e
    # esses saem da lista. Sem isto, um curso sobre Claude Code nao consegue
    # escrever "Claude Code" na tela. Ver o comentario no gerar.py.
    ensina = [t.strip().lower() for t in _gerador.CURSO.get("ensina", []) if t.strip()]
    vis = texto_visivel(html)
    falhas = []
    for termo in VOCAB_INTERNO:
        if any(re.fullmatch(termo, e) for e in ensina):
            continue
        for _, l in linhas_com(vis, termo):
            falhas.append("vocabulário interno {}: {}".format(termo, l))
    return falhas


DIRECAO_DE_CENA = [
    r"pergunte à sala", r"pergunte para a sala", r"a sala responde",
    r"espere o silêncio", r"aguarde o silêncio", r"plano b",
    r"o que apontar", r"roteiro de palco", r"dê um tempo",
    r"circule pela sala", r"anote no quadro", r"projete a tela",
    r"executa ao vivo", r"um grupo por vez", r"minutos por grupo",
    r"apresentem em voz alta", r"lei[ea]m? só até",
]


def g3_direcao_de_cena(rel, html):
    """A página é do aluno, não roteiro do instrutor."""
    if e_vitrine(rel):
        return []
    vis = texto_visivel(html)
    falhas = []
    for termo in DIRECAO_DE_CENA:
        for _, l in linhas_com(vis, termo):
            falhas.append("direção de cena {}: {}".format(termo, l))
    return falhas


def g4_minutagem_fora_da_capa(rel, html):
    """Duração é controle de condução. Só a capa carrega, porque lá é o contrato.

    Quem leva 40 minutos numa aula marcada como de 25 conclui que é lento, e
    para de tentar. A exceção é o cartão de vídeo: a duração de uma GRAVAÇÃO é
    fato do material, do mesmo tipo que "sete páginas".
    """
    if rel in PAGINAS_COM_AGENDA or e_vitrine(rel):
        return []
    vis = texto_visivel(re.sub(r'<a class="fonte".*?</a>', " ", html, flags=re.S))
    # O que é proibido é a duração DO BLOCO DE AULA. "Uma reunião de 20 minutos" e
    # "3 horas por semana" são fato do caso, e apagar isso tiraria do exercício o
    # número que faz ele valer a pena. O gate só acusa quando a duração está na
    # mesma linha que a palavra que nomeia o pedaço do treinamento.
    aula = re.compile(r"\b(aula|bloco|m[óo]dulo|sess[ãa]o|dinâmica|intervalo|"
                      r"exerc[íi]cio|atividade|workshop|treinamento|abertura|"
                      r"encerramento)\b", re.I)
    falhas = []
    for padrao in [r"\d+\s*min\b", r"\b\d+h\d{2}\b", r"\b\d+\s*minutos\b",
                   r"\b\d+\s*horas?\b"]:
        for _, l in linhas_com(vis, padrao):
            if not aula.search(l):
                continue
            falhas.append("duração de aula fora da capa: " + l)
    return falhas


def g5_prompt_tem_os_quatro_paragrafos(rel, html):
    """Todo prompt copiável tem os quatro parágrafos, na ordem.

    O quarto é o que transforma pedido em procedimento: sem "na dúvida", o
    modelo preenche buraco por estimativa e ninguém percebe.
    """
    # Só o prompt que tem botão de copiar. O mesmo bloco em mono também serve para
    # MOSTRAR um arquivo (um regras.md, um trecho de configuração), e arquivo não
    # tem por que ter os quatro parágrafos de um pedido.
    copiaveis = set(re.findall(r'data-alvo="([^"]+)"', html))
    falhas = []
    for i, (classes, ident, corpo) in enumerate(re.findall(
            r'<pre class="prompt-txt([^"]*)"[^>]*id="([^"]+)"[^>]*>(.*?)</pre>',
            html, flags=re.S), 1):
        if ident not in copiaveis:
            continue
        # A EXCECAO QUE O COMENTARIO ACIMA JA DECLARAVA e o codigo nao tinha.
        # Um trecho de CONFIGURACAO copiavel (o perfil da conta, um arquivo de
        # regras) nao e um pedido: ele nao pede nada, descreve quem pergunta.
        # A marca e estrutural, nao caso a caso -- quem quer a dispensa escreve
        # .config na classe, e o gate continua cobrando todo o resto.
        if "config" in classes.split():
            continue
        texto = corpo.replace("&nbsp;", " ")
        for peca in ("O que eu preciso:", "Restrições:", "Na dúvida:"):
            if peca not in texto:
                falhas.append("prompt {} de {}: falta o parágrafo {!r}".format(i, rel, peca))
    return falhas


# =========================================================================
# ESTRUTURA
# =========================================================================

def g6_classe_sem_css(rel, html):
    """Toda classe usada no HTML tem regra no CSS da própria página."""
    definidas = set(re.findall(r"\.([A-Za-z][\w-]*)", css_da_pagina(html)))
    usadas = set()
    for attr in re.findall(r'class="([^"]+)"', html):
        usadas.update(attr.split())
    return ["classe sem CSS: ." + c for c in sorted(usadas - definidas)]


def g7_links_resolvem(rel, html):
    """Todo link relativo aponta para arquivo que existe no disco."""
    base = os.path.dirname(os.path.join(RAIZ, rel))
    falhas = []
    for href in re.findall(r'href="([^"#]+)"', html):
        if href.startswith(("http://", "https://", "mailto:", "data:")):
            continue
        alvo = os.path.normpath(os.path.join(base, href))
        if os.path.isdir(alvo):
            alvo = os.path.join(alvo, "index.html")
        if not os.path.exists(alvo):
            falhas.append("link quebrado: {} (procurei em {})".format(
                href, os.path.relpath(alvo, RAIZ)))
    return falhas


def g8_imagens_existem(rel, html):
    """G7 confere href, e ninguém conferia src.

    O caminho da imagem muda entre a capa (_img/) e a página interna (../_img/),
    que é exatamente o erro que só aparece como quadrado vazio na tela de quem
    abriu.
    """
    base = os.path.dirname(os.path.join(RAIZ, rel))
    falhas = []
    for src in re.findall(r'<img[^>]+src="([^"]+)"', html):
        if src.startswith(("http://", "https://", "data:")):
            continue
        alvo = os.path.normpath(os.path.join(base, src))
        if not os.path.exists(alvo):
            falhas.append("imagem que não existe: {} (procurei em {})".format(
                src, os.path.relpath(alvo, RAIZ)))
    return falhas


def g9_gabarito_fechado(rel, html):
    """O gabarito nunca nasce aberto: ninguém tenta o exercício com a resposta à vista."""
    return ["gabarito nasce aberto: " + m.group(0)[:100]
            for m in re.finditer(r"<details[^>]*>", html)
            if "gabarito" in m.group(0) and re.search(r"\bopen\b", m.group(0))]


def g10_botao_de_copiar_tem_alvo(rel, html):
    """Botão de copiar aponta para um id que existe na mesma página.

    Se o id não bate, o botão não faz nada e ninguém percebe: a página abre, o
    botão aparece, e a falha só existe no clique.
    """
    ids = set(re.findall(r'id="([^"]+)"', html))
    return ["botão de copiar aponta para id inexistente: " + a
            for a in re.findall(r'data-alvo="([^"]+)"', html) if a not in ids]


def g11_numeracao_sequencial(rel, html):
    """Seção e passo numeram de 01 em diante, sem pulo e sem repetição."""
    falhas = []
    for classe, nome in (("secao-n", "seção"), ("passo-n", "passo")):
        rotulos = re.findall(r'<div class="%s">\s*(\d+)\s*</div>' % classe, html)
        if not rotulos:
            continue
        nums = [int(r) for r in rotulos]
        # A seção 00 é a regra que vale para a página inteira, e vem antes da
        # primeira. Quando ela existe, a contagem começa nela.
        inicio = nums[0] if nums and nums[0] == 0 else 1
        if nums != list(range(inicio, inicio + len(nums))):
            falhas.append("a numeração de {} não é sequencial: {}".format(nome, nums))
    return falhas


def g12_secao_tem_funcao_e_titulo(rel, html):
    """Toda seção traz o número, o rótulo de função e o H2.

    O rótulo em cinza diz a FUNÇÃO da seção e o H2 diz o assunto. Sem o rótulo a
    página vira uma pilha de títulos e o aluno perde a anatomia.
    """
    falhas = []
    for m in re.finditer(r'<section class="secao"[^>]*id="([^"]+)"[^>]*>(.{0,700})',
                         html, flags=re.S):
        ident, topo = m.group(1), m.group(2)
        for peca in ('class="secao-n"', 'class="secao-fn"', "<h2>"):
            if peca not in topo:
                falhas.append("seção #{}: falta {}".format(ident, peca))
    return falhas


def g13_tabela_no_envelope(rel, html):
    """Tabela larga fica dentro do envelope que rola sozinho.

    Sem o envelope a página inteira passa a rolar de lado no celular, levando o
    H1 e todo parágrafo junto.
    """
    falhas = []
    for m in re.finditer(r"<table\b", html):
        antes = html[max(0, m.start() - 400):m.start()]
        if antes.rfind('class="tabela') <= antes.rfind("</div>"):
            falhas.append("tabela fora do envelope que rola: "
                          + html[m.start():m.start() + 70].replace("\n", " "))
    return falhas


# =========================================================================
# A QUEBRA DE LINHA · as duas curas, que se parecem e são opostas
# =========================================================================

def g14_cola_aplicada(rel, html):
    """A cola de espaço rígido está aplicada, e nenhum trecho colado estoura.

    A prova é a da régua em papel: rodar o passo do gerador de novo no artefato
    final não pode mudar nada. Passo de geração some em silêncio, e o defeito só
    reaparece na tela de quem abriu.
    """
    falhas = []
    if _gerador.cola_quebra_de_linha(html) != html:
        falhas.append("a cola de quebra de linha não está aplicada: rodar o passo "
                      "do gerador de novo ainda muda o arquivo")
    limite = _gerador.LIMITE_GRUDADO
    # Unidade colada maior que a linha do celular vira rolagem lateral, que é o
    # defeito que esta correção deveria evitar.
    for m in re.finditer(r"(?:[^\s<>]+(?:&nbsp;))+[^\s<>]+", html):
        colado = m.group(0).replace("&nbsp;", " ")
        if "<" in colado or ">" in colado:
            continue
        if len(colado) > limite:
            falhas.append("trecho colado maior que o limite de {}: {!r}".format(
                limite, colado[:60]))
    return falhas


TITULOS = ("h1", "h2", "h3", "h4", "h5", "h6")


def g15_prosa_sem_quebra_equilibrada(rel, html):
    """text-wrap existe só em h1..h4. Nunca em prosa.

    balance e pretty RESERVAM espaço no fim da linha para equilibrar o bloco. Em
    título isso é o efeito desejado; em prosa é a frase que quebra do nada com
    meia linha vazia à direita, que o Rafael reprovou três vezes.

    Medido no site do Longevidade, 8 páginas: 62 quebras precoces com balance na
    prosa, 0 sem ele.
    """
    falhas = []
    for bloco in css_da_pagina(html).split("}"):
        sel, chave, corpo = bloco.partition("{")
        if not chave or "text-wrap" not in corpo:
            continue
        if re.search(r"text-wrap[^:;]*:\s*pretty", corpo):
            falhas.append("{}: text-wrap:pretty também abre folga no fim da linha"
                          .format(rel))
        sel = re.sub(r"/\*.*?\*/", " ", sel, flags=re.S).strip().split("\n")[-1].strip()
        for alvo in [a.strip() for a in sel.split(",") if a.strip()]:
            ultimo = alvo.split()[-1].split(">")[-1].strip()
            if ultimo.lower() not in TITULOS:
                falhas.append("{}: {!r} leva text-wrap e não é título".format(rel, alvo))
    return falhas


# =========================================================================
# A MARCA · trocar de cliente tem de ser trocar um arquivo
# =========================================================================

def _valores_do_css(css):
    """Só o lado direito das declarações. Seletor não entra: #s1 e #fade são id,
    não cor, e olhar o bloco inteiro faria o gate acusar a própria âncora."""
    css = re.sub(r"/\*.*?\*/", " ", css, flags=re.S)
    for bloco in css.split("}"):
        _, chave, corpo = bloco.partition("{")
        if not chave:
            continue
        for decl in corpo.split(";"):
            prop, dois, valor = decl.partition(":")
            if dois:
                yield prop.strip(), valor.strip()


HEX = re.compile(r"#[0-9a-fA-F]{3,8}\b")


def g16_cor_so_no_marca_css(rel, html):
    """Nenhuma cor escrita fora do marca.css.

    É a promessa inteira do template: trocar de cliente é trocar um arquivo.
    Quando um hex vaza para o base.css, ele deixa de ser template e vira o CSS
    daquele cliente, e o vazamento só aparece meses depois, na cor errada de um
    componente que ninguém lembrava que existia.

    A sombra em rgba fica: ela é profundidade, não identidade, e não muda de
    cliente para cliente.
    """
    css = css_da_pagina(html)
    resto = css.replace(MARCA, "", 1)
    if resto == css:
        return ["o marca.css não aparece inteiro no <style> desta página: "
                "o gerador mudou de forma e este gate parou de valer"]
    falhas = []
    for prop, valor in _valores_do_css(resto):
        m = HEX.search(valor)
        if m:
            falhas.append("{}: cor escrita fora do marca.css em {}: {}".format(
                rel, prop, m.group(0)))
    return falhas


def g17_hex_de_oito_digitos(rel, html):
    """Hex de 8 dígitos dentro de SVG não renderiza, e falha em silêncio.

    fill="#26262322" é CSS válido, o XML abre sem erro, o servidor entrega 200
    com o content-type certo e o console fica limpo. A imagem simplesmente não
    aparece. Seis dígitos, e a transparência resolve com opacity.
    """
    return ['{}: hex de 8 dígitos em atributo de SVG: {}'.format(rel, m.group(0))
            for m in re.finditer(r'(?:fill|stroke|stop-color)="#[0-9a-fA-F]{8}"', html)]


# =========================================================================
# AS PEÇAS QUE O ALUNO OPERA
# =========================================================================

CAMPOS_DO_CRIADOR = ["PAPEL", "CONTEXTO", "TAREFA", "FORMATO",
                     "LIMITAÇÕES", "CRITÉRIO DE SUCESSO"]


def g18_criador_completo(rel, html):
    """O criador tem os seis campos, os três botões e o lugar da saída.

    E ele nasce PREENCHIDO: ferramenta que abre vazia é formulário, e formulário
    vazio numa sala faz a pessoa olhar para o lado antes de digitar.
    """
    falhas = []
    for i, (_, bloco) in enumerate(blocos_por_classe(html, "criador"), 1):
        rot = "criador {} de {}".format(i, rel)
        titulos = re.findall(r'data-titulo="([^"]+)"', bloco)
        if titulos != CAMPOS_DO_CRIADOR:
            falhas.append("{}: os campos não são os seis do método, na ordem: {}"
                          .format(rot, titulos))
        for acao in ("exemplo", "limpar", "copiar"):
            if 'data-acao="%s"' % acao not in bloco:
                falhas.append("{}: falta o botão {}".format(rot, acao))
        if 'class="cr-txt"' not in bloco:
            falhas.append(rot + ": falta o lugar onde o prompt monta")
        for m in re.finditer(r'<textarea[^>]*data-titulo="([^"]+)"[^>]*>(.*?)</textarea>',
                             bloco, flags=re.S):
            if not m.group(2).strip():
                falhas.append("{}: o campo {} abre vazio".format(rot, m.group(1)))
        for m in re.finditer(r'<textarea[^>]*id="([^"]+)"', bloco):
            if 'for="%s"' % m.group(1) not in bloco:
                falhas.append("{}: o campo {} não tem label".format(rot, m.group(1)))
    return falhas


def g19_canvas_completo(rel, html):
    """O canvas tem chave própria, chip por campo, o "Ex:" fora da caixa,
    o indicador de rascunho e os dois botões.

    A chave é o que separa o rascunho de um canvas do rascunho de outro no mesmo
    aparelho: dois canvas com a mesma chave se sobrescrevem em silêncio.
    """
    falhas = []
    chaves = []
    for i, (attrs, corpo) in enumerate(blocos_por_classe(html, "canvas"), 1):
        rot = "canvas {} de {}".format(i, rel)
        chave = re.search(r'data-chave="([^"]+)"', attrs)
        if not chave:
            falhas.append(rot + ": sem data-chave, o rascunho colide com o de outro canvas")
        else:
            chaves.append(chave.group(1))
        campos = re.findall(r'<textarea[^>]*id="([^"]+)"', corpo)
        if not campos:
            falhas.append(rot + ": sem campo nenhum")
        if len(campos) != len(re.findall(r'class="cv-n"', corpo)):
            falhas.append(rot + ": tem campo sem o número em chip")
        if len(campos) != len(re.findall(r'class="cv-ex"', corpo)):
            falhas.append(rot + ': tem campo sem o "Ex:" acima da caixa')
        for c in campos:
            if 'for="%s"' % c not in corpo:
                falhas.append("{}: o campo {} não tem label".format(rot, c))
        if 'class="canvas-estado"' not in corpo:
            falhas.append(rot + ": sem o indicador de rascunho salvo")
        for acao in ("linha", "apagar"):
            if 'data-acao="%s"' % acao not in corpo:
                falhas.append("{}: falta o botão {}".format(rot, acao))
    for c in set(chaves):
        if chaves.count(c) > 1:
            falhas.append("{}: dois canvas com a chave {!r}: um apaga o rascunho do outro"
                          .format(rel, c))
    return falhas


def g20_exemplo_imprime_sozinho(rel, html):
    """A página de exemplo pronto vira A4 sem levar o site junto.

    Sem o CSS de impressão, quem imprime leva o cabeçalho, a migalha e os blocos
    de explicação, e o documento perde a cara de documento, que é a única razão
    de ele estar em página própria.
    """
    if 'class="doc ' not in html and 'class="doc"' not in html:
        return []
    falhas = []
    css = css_da_pagina(html)
    if "@media print" not in css:
        falhas.append(rel + ": tem documento e não tem CSS de impressão")
    elif ".topo" not in css.split("@media print", 1)[1][:600]:
        falhas.append(rel + ": o CSS de impressão não esconde a moldura do site")
    if 'class="doc-barra' in html and "so-tela" not in html:
        falhas.append(rel + ": a barra do documento não está marcada como so-tela")
    return falhas


def g21_o_script_nao_esta_quebrado(rel, html):
    """O script da página não tem string partida no meio nem tabulação crua.

    🔴 Esta é a armadilha mais cara do gerador, e ela custou uma rodada inteira.
    A casca é uma string de Python: se ela não for CRUA, a barra-n e a barra-t
    escritas dentro do JavaScript viram quebra de linha e tabulação DE VERDADE no
    arquivo publicado. A página abre normalmente, o layout fica perfeito, e o
    script inteiro morre na primeira linha com "Invalid or unexpected token".

    Nenhum outro gate pega isso: o HTML continua válido, as classes continuam
    todas com CSS, os links continuam resolvendo. O criador de prompt fica lá,
    bonito, e não faz absolutamente nada quando alguém digita.

    Os dois sinais, sem depender de node instalado na máquina:
      1. linha de script com aspa sem par = string partida no meio
      2. tabulação crua dentro do script = uma barra-t que virou tab de verdade
    """
    falhas = []
    for corpo in re.findall(r"<script[^>]*>(.*?)</script>", html, flags=re.S):
        for n, linha in enumerate(corpo.split("\n"), 1):
            if "\t" in linha:
                falhas.append("{}: tabulação crua na linha {} do script: {!r}"
                              .format(rel, n, linha.strip()[:60]))
            # tira o que está escapado antes de contar: \' dentro de string não
            # abre nem fecha nada, e contar ele faz o gate acusar código correto
            limpa = re.sub(r"\\.", "", linha)
            for aspa in ("'", '"'):
                if limpa.count(aspa) % 2:
                    falhas.append("{}: aspa sem par na linha {} do script "
                                  "(string partida no meio): {!r}"
                                  .format(rel, n, linha.strip()[:60]))
    return falhas


QUADRANTES = ("faca", "planeje", "sobrar", "deixe")


def g22_matriz_com_os_quatro_quadrantes(rel, html):
    """A matriz esforço × impacto tem os quatro quadrantes nomeados, e os eixos.

    🔴 O nome do quadrante é o que POSICIONA ele: o CSS declara célula por célula
    e casa pelo nome. Um quadrante sem nome não fica sem cor, fica sem lugar: o
    grid encaixa ele onde sobrar, e a matriz inteira sai em escada, com o eixo X
    no meio do nada.

    O defeito é discreto o bastante para passar por revisão de código e por olho
    em página aberta, porque cada caixa continua bonita. Já aconteceu.
    """
    falhas = []
    for i, (_, corpo) in enumerate(blocos_por_classe(html, "priorizar"), 1):
        rot = "matriz {} de {}".format(i, rel)
        nomes = []
        classes = re.findall(r'<div class="quad([^"]*)"', corpo)
        for c in classes:
            nomes.extend([x for x in c.split() if x in QUADRANTES])
        if sorted(nomes) != sorted(QUADRANTES):
            falhas.append("{}: os quadrantes deviam ser {} e são {}".format(
                rot, list(QUADRANTES), nomes))
        for peca, nome in (("pri-esf-baixo", "o rótulo da coluna de esforço baixo"),
                           ("pri-esf-alto", "o rótulo da coluna de esforço alto"),
                           ("pri-eixo-y", "o eixo de impacto"),
                           ("pri-eixo-x", "o eixo de esforço")):
            if peca not in corpo:
                falhas.append("{}: falta {} ({})".format(rot, nome, peca))
    return falhas


def g23_pergunta_de_grupo_tem_destrave(rel, html):
    """Ou todas as perguntas do bloco têm destrave, ou nenhuma tem.

    O grupo que recebe destrave numa pergunta e não na seguinte conclui que a
    segunda é mais fácil, e responde mais raso. E toda pergunta com destrave diz
    o que NÃO conta como resposta: sem isso, metade da sala responde "numa pasta
    compartilhada" e acha que respondeu.
    """
    falhas = []

    # 🔴 O EXERCICIO TAMBEM. Ate 28/08 este gate so varria .perguntas, e entao
    # um exercicio de quatro passos sem nenhum destrave passava nas 864
    # checagens com zero achado. Foi o defeito que o autor do curso achou DUAS
    # vezes no piloto IEL, apontando a pagina de origem, enquanto a suite dizia
    # "passou". No exercicio a unidade e o proprio .passo, nao a .pergunta:
    # ou todos os passos destravam, ou nenhum destrava. O passo que destrava ao
    # lado do que nao destrava ensina que o segundo e mais facil.
    passos = [c for _, c in blocos_por_classe(html, "passo")]
    if passos:
        com_d = [c for c in passos if 'class="destrave"' in c]
        # NAO e "tudo ou nada" aqui: "abra uma conversa nova" nao tem o que
        # destravar. O que nao pode e o exercicio inteiro sem nenhum -- foi
        # assim que as 5 aulas de nivelamento do piloto IEL sairam, 17 passos
        # e zero destrave, e a suite deu "passou".
        if not com_d:
            falhas.append("{}: {} passos e nenhum destrave. Todo passo que pede "
                          "o aluno a produzir texto proprio precisa de um"
                          .format(rel, len(passos)))
        for corpo in com_d:
            if 'class="destrave-nao"' not in corpo:
                falhas.append("{}: um passo com destrave nao diz o que NAO conta "
                              "como resposta".format(rel))

    for i, (_, bloco) in enumerate(blocos_por_classe(html, "perguntas"), 1):
        perguntas = blocos_por_classe(bloco, "pergunta")
        if not perguntas:
            continue
        com = [c for _, c in perguntas if 'class="destrave"' in c]
        if com and len(com) != len(perguntas):
            falhas.append("{}: bloco {} tem {} de {} perguntas com destrave: "
                          "ou todas têm, ou o padrão quebrou"
                          .format(rel, i, len(com), len(perguntas)))
        for corpo in com:
            if 'class="destrave-nao"' not in corpo:
                falhas.append("{}: bloco {}: uma pergunta não diz o que NÃO conta "
                              "como resposta".format(rel, i))
    return falhas


def g24_os_dois_loops_comecam_iguais(rel, html):
    """No par aberto × fechado, a segunda cadeia começa igual à primeira.

    🔴 A figura ensina por diferença: a pessoa compara as duas cadeias e vê o que
    sobra. Se a segunda reescreve o primeiro passo, o olho passa a comparar texto
    em vez de comparar estrutura, e a figura perde o argumento inteiro.

    O fechado também precisa ser MAIOR: dois loops do mesmo tamanho não mostram
    diferença nenhuma.
    """
    falhas = []
    for i, (_, bloco) in enumerate(blocos_por_classe(html, "loops"), 1):
        cadeias = []
        for m in re.finditer(r'<div class="loop\b[^"]*"[^>]*>(.*?)(?=<div class="loop\b|$)',
                             bloco, flags=re.S):
            nos = [re.sub(r"\s+", " ", n).strip() for n in
                   re.findall(r'<span class="loop-no">(.*?)</span>', m.group(1), flags=re.S)]
            if nos:
                cadeias.append(nos)
        if len(cadeias) != 2:
            falhas.append("{}: o par {} tem {} cadeia(s): a figura só existe com duas"
                          .format(rel, i, len(cadeias)))
            continue
        aberto, fechado = cadeias
        if fechado[:len(aberto)] != aberto:
            falhas.append("{}: o par {} não começa igual: aberto {} · fechado {}"
                          .format(rel, i, aberto, fechado[:len(aberto)]))
        if len(fechado) <= len(aberto):
            falhas.append("{}: o par {}: o loop fechado não é maior que o aberto"
                          .format(rel, i))
    return falhas


def g25_o_cem_bate_com_a_legenda(rel, html):
    """A parede tem cem quadrados, e os acesos batem com o número escrito.

    🔴 Ninguém confere contando cem quadradinhos. A figura pode passar meses no
    ar dizendo "12" com dezoito acesos, e o defeito não parece defeito: parece
    uma parede de quadrados.

    O gate compara três coisas que têm de concordar: o data-acesos declarado, os
    <i class="aceso"> que o gerador produziu, e o primeiro número da legenda.
    """
    falhas = []
    for i, (attrs, corpo) in enumerate(blocos_por_classe(html, "cem-grade"), 1):
        rot = "cem {} de {}".format(i, rel)
        m = re.search(r'data-acesos="(\d+)"', attrs)
        if not m:
            falhas.append(rot + ": sem data-acesos, a grade não pode ser gerada")
            continue
        declarado = int(m.group(1))
        total = len(re.findall(r'<i class="cem-p', corpo))
        acesos = len(re.findall(r'<i class="cem-p aceso"', corpo))
        if total != 100:
            falhas.append("{}: a parede tem {} quadrados, e não cem: o passo do "
                          "gerador não rodou".format(rot, total))
        if acesos != declarado:
            falhas.append("{}: declara {} acesos e desenha {}".format(
                rot, declarado, acesos))
    # a legenda é texto do autor, e é ela que a pessoa lê
    for i, (_, bloco) in enumerate(blocos_por_classe(html, "cem"), 1):
        m = re.search(r'data-acesos="(\d+)"', bloco)
        leg = re.search(r'<i class="aceso"></i>\s*(?:&nbsp;|\s)*(\d+)', bloco)
        if m and leg and int(m.group(1)) != int(leg.group(1)):
            falhas.append("{}: cem {}: a grade acende {} e a legenda diz {}".format(
                rel, i, m.group(1), leg.group(1)))
    return falhas


def g26_colunas_com_tres_listas_iguais(rel, html):
    """Valores, barras e rótulos: três listas do mesmo tamanho, igual a --n.

    🔴 Elas alinham porque compartilham o mesmo grid. Se uma tiver um item a
    menos, o rótulo desliza para a barra do vizinho e o gráfico passa a afirmar
    outra coisa, sem nada parecer quebrado na tela.
    """
    falhas = []
    for i, (attrs, bloco) in enumerate(blocos_por_classe(html, "colunas"), 1):
        rot = "colunas {} de {}".format(i, rel)
        n = re.search(r"--n:\s*(\d+)", attrs)
        if not n:
            falhas.append(rot + ": sem --n, as três listas não têm como alinhar")
            continue
        n = int(n.group(1))
        tamanhos = {}
        for classe in ("col-vs", "col-bs", "col-rs"):
            achados = blocos_por_classe(bloco, classe)
            if not achados:
                falhas.append("{}: falta a lista .{}".format(rot, classe))
                continue
            tamanhos[classe] = len(re.findall(r"<span", achados[0][1]))
        if tamanhos and set(tamanhos.values()) != {n}:
            falhas.append("{}: --n é {} e as listas têm {}".format(rot, n, tamanhos))
        # barra sem altura fica com zero e desaparece sem erro nenhum
        barras = blocos_por_classe(bloco, "col-bs")
        if barras:
            spans = re.findall(r"<span([^>]*)>", barras[0][1])
            sem_h = [a for a in spans if "--h:" not in a]
            if sem_h:
                falhas.append("{}: {} barra(s) sem --h: elas somem com altura zero"
                              .format(rot, len(sem_h)))
    return falhas


def g27_grade_de_n_colunas_pode_encolher(rel, html):
    """Grade com número de colunas variável usa minmax(0,1fr), nunca 1fr puro.

    🔴 `1fr` tem `min-width:auto`: a coluna nunca fica menor que a palavra mais
    comprida dentro dela. Num gráfico de sete colunas, um rótulo como "Entregue,
    negociando" estica a faixa inteira, e a PÁGINA passa a rolar de lado, levando
    o H1 e todo parágrafo junto.

    Aconteceu: 101px de rolagem lateral na vitrine, com todos os outros gates
    passando. Nenhum deles olhava para isso, e o defeito só existe no celular.
    """
    falhas = []
    css = re.sub(r"/\*.*?\*/", " ", css_da_pagina(html), flags=re.S)
    for bloco in css.split("}"):
        sel, chave, corpo = bloco.partition("{")
        if not chave:
            continue
        for m in re.finditer(r"repeat\(\s*var\(--[\w-]+\s*\)\s*,\s*([^)]*?)\s*\)", corpo):
            if "minmax(0" not in m.group(0):
                falhas.append("{}: {!r} monta a grade com {} em vez de minmax(0,1fr): "
                              "a coluna trava na largura da palavra mais comprida"
                              .format(rel, sel.strip().split("\n")[-1].strip()[:60],
                                      m.group(1)))
    return falhas


def g28_faixa_cabe_na_regua(rel, html):
    """Em raia e gantt, nenhuma faixa começa depois de terminar nem passa do fim.

    🔴 `grid-column: var(--de) / calc(var(--ate) + 1)` com --ate maior que --n
    coloca a barra numa coluna implícita, FORA do quadro. A tarefa desaparece do
    plano e nada na tela parece quebrado: sobra um espaço em branco onde ela
    deveria estar, e espaço em branco num gantt lê como "nessa semana não tem
    nada".
    """
    falhas = []
    for classe, faixa in (("raias", "ra-p"), ("gantt", "gt-b")):
        for i, (attrs, corpo) in enumerate(blocos_por_classe(html, classe), 1):
            rot = "{} {} de {}".format(classe, i, rel)
            n = re.search(r"--n:\s*(\d+)", attrs)
            if not n:
                falhas.append(rot + ": sem --n, a régua não tem tamanho")
                continue
            n = int(n.group(1))
            for m in re.finditer(r'<div class="%s[^"]*"[^>]*style="([^"]*)"' % faixa, corpo):
                est = m.group(1)
                de = re.search(r"--de:\s*(\d+)", est)
                ate = re.search(r"--ate:\s*(\d+)", est)
                if not de or not ate:
                    falhas.append(rot + ": faixa sem --de ou --ate")
                    continue
                de, ate = int(de.group(1)), int(ate.group(1))
                if de < 1 or ate > n:
                    falhas.append("{}: faixa de {} a {} não cabe numa régua de {}"
                                  .format(rot, de, ate, n))
                elif de > ate:
                    falhas.append("{}: faixa começa em {} e termina em {}"
                                  .format(rot, de, ate))
            # a régua do cabeçalho tem de ter o mesmo tanto de colunas
            for cab in ("ra-etapas", "gt-regua"):
                for _, r in blocos_por_classe(corpo, cab):
                    marcas = len(re.findall(r"<span", r))
                    if marcas != n:
                        falhas.append("{}: --n é {} e a régua tem {} marcas"
                                      .format(rot, n, marcas))
    return falhas


def g29_radar_bate_eixo_com_valor(rel, html):
    """No radar, cada série tem um valor por eixo, e todo valor vai de 0 a 100.

    🔴 Um valor a mais ou a menos não dá erro: o polígono simplesmente fecha em
    outro lugar, com a mesma cara de gráfico correto. E valor acima de 100
    desenha para fora do último anel, o que faz o eixo parecer cheio quando ele
    está estourado.
    """
    falhas = []
    for i, (attrs, corpo) in enumerate(blocos_por_classe(html, "radar"), 1):
        rot = "radar {} de {}".format(i, rel)
        eixos = re.search(r'data-eixos="([^"]*)"', attrs)
        if not eixos:
            falhas.append(rot + ": sem data-eixos")
            continue
        nomes = [e.strip() for e in eixos.group(1).split("|") if e.strip()]
        if not 3 <= len(nomes) <= 6:
            falhas.append("{}: {} eixos. Com dois não é radar, com sete o rótulo "
                          "não cabe".format(rot, len(nomes)))
        for campo in ("data-valores", "data-valores-b"):
            m = re.search(r'%s="([^"]*)"' % campo, attrs)
            if not m:
                continue
            vs = [v.strip() for v in m.group(1).split(",") if v.strip()]
            if len(vs) != len(nomes):
                falhas.append("{}: {} tem {} valores para {} eixos".format(
                    rot, campo, len(vs), len(nomes)))
            for v in vs:
                if not v.replace(".", "", 1).isdigit() or not 0 <= float(v) <= 100:
                    falhas.append("{}: valor fora de 0 a 100 em {}: {}".format(
                        rot, campo, v))
        if "<svg" not in corpo:
            falhas.append(rot + ": o passo do gerador não desenhou o polígono")
            continue
        # 🔴 O <svg> RECORTA o que passa da moldura, e recortar é o comportamento
        # normal dele: o rótulo simplesmente some, sem erro, sem console, sem
        # nada torto na página. "Padrão do formato" saiu pela borda e o desenho
        # continuou parecendo certo.
        vb = re.search(r'viewBox="([-\d. ]+)"', corpo)
        if not vb:
            falhas.append(rot + ": o svg não declara viewBox")
            continue
        x0, _, larg, _ = [float(x) for x in vb.group(1).split()]
        for m in re.finditer(r'<text x="([-\d.]+)"[^>]*text-anchor="(\w+)"[^>]*>([^<]*)</text>',
                             corpo):
            x, ancora, txt = float(m.group(1)), m.group(2), m.group(3)
            lw = _gerador.largura_estimada(txt)
            esq = x if ancora == "start" else (x - lw if ancora == "end" else x - lw / 2)
            if esq < x0 or esq + lw > x0 + larg:
                falhas.append("{}: o rótulo {!r} sai da moldura e o svg recorta ele"
                              .format(rot, txt))
    return falhas


def g30_mapa_de_area_soma_cem(rel, html):
    """As linhas do mapa somam 100, e as células de cada linha também.

    🔴 Se não somarem, o desenho continua bonito e as proporções passam a ser
    inventadas. É o pior tipo de defeito de figura: ela vira um argumento falso
    com aparência de medida.
    """
    falhas = []
    for i, (_, bloco) in enumerate(blocos_por_classe(html, "mapa-area"), 1):
        rot = "mapa {} de {}".format(i, rel)
        linhas = blocos_por_classe(bloco, "ma-linha")
        alturas = []
        for j, (attrs, corpo) in enumerate(linhas, 1):
            h = re.search(r"--h:\s*([\d.]+)", attrs)
            if not h:
                falhas.append("{}: linha {} sem --h".format(rot, j))
                continue
            alturas.append(float(h.group(1)))
            larguras = [float(x) for x in re.findall(r"--w:\s*([\d.]+)", corpo)]
            if larguras and abs(sum(larguras) - 100) > 0.5:
                falhas.append("{}: as células da linha {} somam {:g}, não 100"
                              .format(rot, j, sum(larguras)))
        if alturas and abs(sum(alturas) - 100) > 0.5:
            falhas.append("{}: as linhas somam {:g}, não 100".format(rot, sum(alturas)))
    return falhas


# Extensão de arquivo escrita como `.xlsx` parece nome de classe para o regex e
# não é. A lista é fechada de propósito: crescer ela por conveniência é como um
# gate ganha exceção permanente e deixa de ser gate.
NAO_SAO_CLASSES = {"xlsx", "docx", "csv", "txt", "html", "css", "md", "py", "json"}


def g31_o_roteiro_cita_classe_que_existe(rel, html):
    """COMO-EXECUTAR.md só cita classe que existe mesmo no base.css.

    🔴 O roteiro é o arquivo que um Claude novo lê para saber o que usar. Se ele
    cita `.tabela-comparativa` e essa classe nunca existiu, o próximo material
    nasce com uma classe morta no HTML: a página abre, o bloco aparece sem
    estilo nenhum, e parece descuido de escrita em vez de referência errada.

    Roda uma vez só. O roteiro não é página publicada, e auditar o mesmo arquivo
    quatro vezes multiplicaria por quatro cada achado.
    """
    if rel != "index.html":
        return []
    caminho = os.path.join(RAIZ, "COMO-EXECUTAR.md")
    if not os.path.exists(caminho):
        return ["COMO-EXECUTAR.md não existe: o roteiro é a porta de entrada do padrão"]
    doc = io.open(caminho, encoding="utf-8").read()
    # Contra o CSS QUE A PÁGINA ENTREGA, não contra o base.css do disco. É o que
    # o aluno recebe, e é o que o defeito injetado consegue mexer.
    css = re.sub(r"/\*.*?\*/", " ", css_da_pagina(html), flags=re.S)
    existem = set(re.findall(r"\.([a-z][\w-]*)", css))
    falhas = []
    for c in sorted(set(re.findall(r"`\.([a-z][\w-]*)", doc))):
        if c not in existem and c not in NAO_SAO_CLASSES:
            falhas.append("o roteiro cita .{} e essa classe não existe no base.css".format(c))
    return falhas


def _manifesto():
    caminho = os.path.join(RAIZ, "_build", "insumos.json")
    if not os.path.exists(caminho):
        return {}
    return json.loads(io.open(caminho, encoding="utf-8").read())


def _abre_mesmo(caminho, info):
    """Abre o insumo de verdade e confere que as abas e o volume batem com o
    manifesto. Devolve a mensagem do problema, ou None."""
    if not caminho.lower().endswith(".xlsx"):
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, read_only=True)
    except Exception as e:
        return "{}: {}".format(type(e).__name__, str(e)[:90])
    try:
        if wb.sheetnames != info["abas"]:
            return "as abas são {} e o manifesto diz {}".format(
                wb.sheetnames, info["abas"])
        ws = wb[info["abas"][1]] if len(info["abas"]) > 1 else wb.worksheets[0]
        dados = ws.max_row - info["cabecalho_na_linha"]
        if dados != info["linhas"]:
            return "a aba tem {} linhas de dado e o manifesto diz {}".format(
                dados, info["linhas"])
    finally:
        wb.close()
    return None


def g32_a_pagina_bate_com_o_insumo(rel, html):
    """A página de caso só afirma sobre o insumo o que o arquivo tem de verdade.

    🔴 Este é o acoplamento que quebra em silêncio. O prompt do passo 4 cita as
    colunas do arquivo; se um mudar e o outro não, o aluno cola o prompt e a IA
    responde sobre uma coluna que não existe. A resposta sai plausível e errada,
    na frente da sala, e nada na página parece quebrado.

    Confere três coisas contra o manifesto que o gerador produz:
      1. o arquivo oferecido existe no disco
      2. o número de linhas e de abas que a página anuncia é o real
      3. toda coluna citada em Maiúscula dentro do prompt existe em alguma aba
    """
    manifesto = _manifesto()
    if not manifesto:
        return []
    falhas = []
    for slug, info in manifesto.items():
        # Uma aula pode ter mais de uma demonstracao: "caso" aceita string
        # (um caso) ou lista (varios). Antes so a string passava, e a aula com
        # duas demonstracoes quebrava o manifesto.
        casos = info["caso"]
        if isinstance(casos, str):
            casos = [casos]
        if rel not in [c + "/index.html" for c in casos]:
            continue
        rot = "{} de {}".format(info["arquivo"], rel)

        # 1 · o arquivo existe mesmo
        if info["arquivo"] not in html:
            falhas.append(rot + ": a página não oferece o arquivo do manifesto")
            continue
        no_disco = os.path.join(RAIZ, "_arquivos", info["arquivo"])
        if not os.path.exists(no_disco):
            falhas.append(rot + ": oferecido para download e não existe no disco. "
                                "Rode python3 _build/insumo.py")
        else:
            # 🔴 Existir não é abrir. Um passo que regravava o zip para congelar
            # o carimbo quebrou o XML interno, e o .xlsx parou de abrir: os 34
            # gates continuaram passando, porque nenhum tentava ABRIR o arquivo.
            # O aluno é que descobriria, na sala.
            erro = _abre_mesmo(no_disco, info)
            if erro:
                falhas.append("{}: o arquivo está no disco e NÃO ABRE: {}".format(rot, erro))

        vis = texto_visivel(html)

        # 2 · os números anunciados
        linhas_fmt = "{:,}".format(info["linhas"]).replace(",", ".")
        if linhas_fmt not in vis and str(info["linhas"]) not in vis:
            falhas.append("{}: o arquivo tem {} linhas e a página não diz isso"
                          .format(rot, linhas_fmt))
        for m in re.finditer(r"(\d+)\s*abas", vis):
            if int(m.group(1)) != len(info["abas"]):
                falhas.append("{}: a página anuncia {} abas e o arquivo tem {}"
                              .format(rot, m.group(1), len(info["abas"])))

        # 3 · as colunas citadas no prompt existem em alguma aba
        #
        # 🔴 blocos_por_classe() só varre <div>, e .prompt-txt é um <pre>. Este
        # gate nasceu cego por causa disso: zero achado e zero prova, com a
        # aparência de gate que passa.
        reais = {c.strip().lower() for cols in info["colunas"].values() for c in cols}
        for m in re.finditer(r'<pre class="[^"]*\bprompt-txt\b[^"]*"[^>]*>(.*?)</pre>',
                             html, flags=re.S):
            texto = re.sub(r"<[^>]+>", " ", m.group(1)).replace("&nbsp;", " ")
            # o nome pode ter mais de uma palavra ("Valor Total"): capturar só a
            # primeira faria o gate acusar coluna real como inexistente
            for c in re.finditer(r"\bcoluna\s+([A-Z][\wÀ-ÿ]*(?:\s+[A-Z][\wÀ-ÿ]*)*)",
                                 texto):
                nome = c.group(1).strip()
                if nome.lower() not in reais:
                    falhas.append("{}: o prompt cita a coluna {!r} e ela não existe "
                                  "em aba nenhuma do arquivo".format(rot, nome))
    return falhas


# 🔴 A lista de frases proibidas FALHA, e falhou de verdade: o gate tinha
# "leia em voz alta" e o texto dizia "leio em voz alta". Passou.
#
# A regra que substitui: no bloco que narra a demonstração, VERBO EM PRIMEIRA
# PESSOA DO SINGULAR reprova. Quem opera o teclado não aparece na página do
# aluno; o que aparece é o que acontece na tela.
#
# A correção inversa importa igual: "vocês" sozinho NÃO é defeito. "O dia a dia
# de vocês" trata a turma como profissionais, e é dos melhores trechos que ele
# já aprovou. O defeito é a turma como PLATEIA: "pergunto para vocês",
# "na frente de vocês".
# 🔴 Só verbo que NÃO é substantivo comum. A primeira versão tinha "passo",
# "projeto", "jogo", "colo", "rodo", "conto" e "começo": 39 falsos positivos numa
# página em que "passo 4" é o nome do bloco. É a mesma armadilha de "alcançar" e
# "delegar", e gate com falso positivo vira gate com exceção, que deixa de ser
# gate.
EU_INSTRUTOR = re.compile(
    r"(?<![\wÀ-ÿ])(abro|mostro|paro|pergunto|explico|falo|digito"
    r"|escrevo|leio|aponto|clico|peço|repito|volto|termino|encerro"
    r"|apresento|reforço|destaco)(?![\wÀ-ÿ])", re.I)

# 🔴 Dentro de um prompt de exemplo, a primeira pessoa é do ALUNO, não do
# instrutor: "# QUEM EU SOU · Supervisor de 12 lojas. Apresento para a diretoria"
# é exatamente o que o padrão manda escrever ali. Estes blocos saem da varredura.
VOZ_DO_ALUNO = re.compile(
    r'<pre class="[^"]*\b(?:prompt-txt|cr-txt)\b[^"]*"[^>]*>.*?</pre>'
    r"|<textarea\b.*?</textarea>"
    # O .cv-rot e ROTULO DE CANVAS: texto que o aluno preenche, na voz dele.
    # Sem esta linha o gate acusava "escrevo" dentro de um rotulo e mandava
    # consertar o que estava certo. Achado no piloto IEL.
    r'|<[^>]*class="[^"]*\bcv-rot\b[^"]*"[^>]*>.*?</[a-zA-Z]+>',
    re.S | re.I)

TURMA_COMO_PLATEIA = re.compile(
    r"(?<![\wÀ-ÿ])(para vocês|pra vocês|na frente de vocês|vocês veem|vocês vão ver"
    r"|mostro a vocês|pergunto a vocês)(?![\wÀ-ÿ])", re.I)


def g33_narrativa_sem_instrutor(rel, html):
    """Na página do aluno, o instrutor não aparece na primeira pessoa.

    🔴 Este gate existe porque a LISTA DE FRASES não pega. Ela tinha "leia em
    voz alta" e o texto dizia "leio em voz alta": a mesma direção de cena, uma
    letra de diferença, e passou. Pessoa gramatical pega; palavra não.

    Ele não reprova "vocês": tratar a turma como profissional é bom. Reprova a
    turma como plateia.
    """
    falhas = []
    vis = texto_visivel(VOZ_DO_ALUNO.sub(" ", html))
    for n, linha in enumerate(vis.split("\n"), 1):
        for m in EU_INSTRUTOR.finditer(linha):
            falhas.append("{}: verbo do instrutor em 1ª pessoa ({!r}): {}"
                          .format(rel, m.group(1), linha[:110]))
        for m in TURMA_COMO_PLATEIA.finditer(linha):
            falhas.append("{}: a turma tratada como plateia ({!r}): {}"
                          .format(rel, m.group(1), linha[:110]))
    return falhas


def g34_fr_host_tem_as_frases_marcadas(rel, html):
    """Todo bloco .fr-host teve as frases cortadas pelo gerador.

    🔴 O passo do gerador some em silêncio: basta alguém editar o HTML publicado
    à mão. E o defeito não aparece na tela larga por acaso, ele aparece como a
    quarta reclamação da mesma quebra de linha.
    """
    falhas = []
    for i, (attrs, corpo) in enumerate(blocos_por_classe(html, "fr-host"), 1):
        visivel = re.sub(r"<[^>]+>", " ", corpo).replace("&nbsp;", " ")
        # bloco de uma frase só não precisa de marca
        frases = len(re.findall(r"[.!?](?:\s)+(?=[A-ZÀ-Ý])", visivel))
        marcadas = corpo.count('<span class="fr">')
        if frases >= 1 and marcadas == 0:
            falhas.append("{}: .fr-host {} tem {} frases e nenhuma marcada: "
                          "rode python3 _build/gerar.py".format(rel, i, frases + 1))
    return falhas


def g35_breakout_nao_sobrevive_a_shorthand(rel, html):
    """Componente que anda com .solta nao pode declarar margin em shorthand.

    O breakout do padrao e `margin-left:50%` mais `transform:translateX(-50%)`.
    Um `margin:22px 0` na regra do proprio componente zera o margin-left e NAO
    toca o transform: o bloco continua deslocado meia largura para a esquerda,
    agora sem margem nenhuma para compensar, e sai pela borda da tela.

    🔴 POR QUE ESTE GATE EXISTE, medido em 26/08 na .aulas-lista:
       o cartao de aula nascia em -311px com 1265px de viewport, cortado pela
       esquerda, SEM gerar rolagem horizontal (transform nao cria area de
       rolagem) e SEM nenhum gate acusar. Passou em 238 checagens. So apareceu
       porque o Rafael abriu o site e olhou.

    Defeito que nao produz sintoma mecanico e o mais caro que existe: ele
    atravessa a revisao inteira e chega no cliente.
    """
    css = css_da_pagina(html)

    # 🔴 SO VALE PARA QUEM VEM DEPOIS DE .solta NO ARQUIVO.
    # Especificidade igual (uma classe contra uma classe): quem ganha e a
    # ultima escrita. Regra ANTES de .solta perde e e inofensiva -- e o caso
    # de .contraste, .converge, .demo, .tabela e mais trinta e cinco, que
    # declaram margin em shorthand ha meses e funcionam. Reprovar as trinta e
    # cinco por um defeito que so existe depois da linha do .solta seria
    # gate que grita sem ter razao, e gate que grita sem razao e desligado.
    # A regra BASE do .solta, e nao um `.com-trilha .solta` de override: por
    # isso o ^ de inicio de linha. Sem ele a ancora cai no primeiro override,
    # 1800 linhas acima, e o gate reprova o arquivo inteiro.
    m = re.search(r"(?m)^\.solta\s*\{", css)
    if not m:
        return ["o .solta sumiu do CSS desta pagina: este gate parou de valer"]
    corte = m.start()

    vizinhas = set()
    for attr in re.findall(r'class="([^"]*\bsolta\b[^"]*)"', html):
        for c in attr.split():
            if c != "solta":
                vizinhas.add(c)
    falhas = []
    for c in sorted(vizinhas):
        for reg in re.finditer(r"(?<![\w-])\.%s\s*\{([^}]*)\}" % re.escape(c), css):
            if reg.start() < corte:
                continue
            if re.search(r"(^|;)\s*margin\s*:", reg.group(1)):
                falhas.append(
                    "a classe .{} anda com .solta, e a regra dela vem DEPOIS do "
                    ".solta no arquivo declarando margin em shorthand. O shorthand "
                    "zera o margin-left do breakout e o transform sobrevive "
                    "sozinho: use margin-top e margin-bottom.".format(c))
    return falhas


def g36_a_mesa_fecha_cem(rel, html):
    """Toda linha da .mesa soma exatamente 100%.

    A mesa ensina uma conta: quanto de um espaco finito cada coisa ocupa.
    Barra que nao fecha 100 mente sobre a propria conta que esta ensinando, e
    o aluno nao tem como perceber -- ele nao vai medir a barra com regua.
    """
    falhas = []
    pedacos = html.split('<div class="mesa-linha">')[1:]
    for i, pedaco in enumerate(pedacos, 1):
        larguras = [float(x) for x in re.findall(
            r'class="mesa-faixa[^"]*"\s+style="width:([\d.]+)%"', pedaco)]
        if not larguras:
            continue
        soma = round(sum(larguras), 2)
        if abs(soma - 100) > 0.01:
            falhas.append("a mesa {} de {} soma {:g}%, e nao 100%".format(i, rel, soma))
    return falhas


# Filhos diretos legitimos do <main>. Levantado das 6 paginas do padrao em
# 28/08. Peca de CONTEUDO nunca entra aqui: ela mora dentro de um destes.
PRIMEIRO_NIVEL = {
    "secao", "passo", "heroi", "aviso", "nesta-aula",
    "checagem", "fecho", "gancho", "rodape-nav", "doc", "doc-barra",
    # o divisor da aula com dois comprimentos, e a nota dele
    "ate-aqui", "ate-aqui-nota",
}


def g37_as_divs_fecham(rel, html):
    """Cada <div> aberta fecha.

    🔴 Nasceu de defeito real no piloto IEL: um <div> do .grao copiado da
    vitrine sem a linha de fechamento. Toda a secao 02 caiu dentro de um grid
    de duas colunas e a pagina passou a rolar na horizontal. Os 36 gates deram
    zero achado e quem pegou foi o navegador.
    """
    abre = len(re.findall(r"<div\b", html))
    fecha = len(re.findall(r"</div\s*>", html))
    if abre != fecha:
        return ["{}: {} <div> aberta(s) para {} fechada(s), diferenca de {}"
                .format(rel, abre, fecha, abs(abre - fecha))]
    return []


def g38_todo_bloco_mora_em_envelope(rel, html):
    """Filho direto do <main> so pode ser envelope, nunca peca de conteudo.

    🔴 Nasceu de defeito real no piloto IEL: um bloco inteiro emendado DEPOIS
    do </section> em vez de antes. O HTML existia, o balanceamento fechava, os
    36 gates deram zero, e o bloco nao estava dentro de secao nenhuma -- entao
    nascia sem a margem, sem a numeracao e fora do sumario.
    """
    m = re.search(r"<main\b[^>]*>", html)
    if not m:
        return []
    falhas, i, prof = [], m.end(), 1
    while i < len(html) and prof:
        tag = re.search(r"<(/?)([a-zA-Z]\w*)([^>]*?)>", html[i:])
        if not tag:
            break
        if prof == 1 and not tag.group(1):
            cls = re.search(r'class="([^"]*)"', tag.group(3))
            nomes = cls.group(1).split() if cls else []
            if not (set(nomes) & PRIMEIRO_NIVEL):
                falhas.append('{}: <{} class="{}"> e filho direto do <main>: '
                              "todo bloco mora dentro de um envelope"
                              .format(rel, tag.group(2), " ".join(nomes)))
        i += tag.end()
        if tag.group(2).lower() in ("br", "img", "input", "meta", "link", "hr", "source"):
            continue
        prof += -1 if tag.group(1) else 1
    return falhas


def _e_aula(html):
    """Pagina de AULA: oito secoes e pelo menos um conceito.

    A ASSINATURA E A PROPRIA ANATOMIA. A aula tem exatamente 8 .secao, porque
    a anatomia e fixa; a vitrine tem 48, o modulo tem 2, a pagina de caso e o
    exemplo tem zero. Medido nas 6 paginas do padrao e nas 9 aulas do piloto
    IEL em 28/08: todas as nove tinham 8, sem excecao.

    Nao depende de NOME de pagina (foi o que deixou tres gates cegos no
    piloto) nem de campo novo no gerador. Se a anatomia mudar de tamanho,
    este numero muda junto, e e por isso que ele mora aqui e nao espalhado.
    """
    secoes = len(re.findall(r'<section class="(?:[^"]* )?secao(?: [^"]*)?"', html))
    if secoes != 8:
        return None
    con = re.findall(r'<[^>]*class="(?:[^"]* )?conceito(?: [^"]*)?"', html)
    pas = re.findall(r'<[^>]*class="(?:[^"]* )?passo(?: [^"]*)?"', html)
    return (len(con), len(pas)) if con else None


def g39_uma_aula_um_conceito(rel, html):
    """🔴 UMA AULA, UM CONCEITO. Dois conceitos sao duas aulas.

    Ate 28/08 a anatomia dizia "o conceito DIVIDIDO se for mais de um", e o
    resultado medido no piloto IEL foi: 9 de 9 aulas com dois conceitos
    empilhados na secao 02 e a primeira pratica so na secao 05. A regra 1 do
    roteiro -- conceito, imagem, pratica, proximo conceito -- estava escrita e
    era inaplicavel, porque duas voltas do ciclo nao cabem em oito secoes.

    A cura nao e reordenar: e cortar. Com um conceito por aula o ciclo fecha
    sozinho, a aula encolhe, e o encontro passa a ter MAIS voltas de pratica,
    nao menos -- que era o que faltava quando a sala ficou parada em 27/08.
    """
    m = _e_aula(html)
    if not m:
        return []
    conceitos, _ = m
    if conceitos > 1:
        return ["{}: {} conceitos numa aula so. Uma aula, um conceito -- o "
                "segundo vira outra aula".format(rel, conceitos)]
    return []


def g41_o_conceito_vem_com_imagem(rel, html):
    """🔴 Todo conceito nasce com a analogia junto. A imagem nao vem depois.

    "As pessoas nao entendem o que voce fala, mas o que elas veem ou o que elas
    sentem quando voce fala" -- e "a imagem vem primeiro, para depois processar
    o entendimento". Da formula AIDEN, de Adriano de Marqui.

    Medido no piloto IEL, aula n1: a secao de conceito tinha 632 palavras e a
    primeira figura so chegava na secao seguinte. Seiscentas palavras de
    abstracao antes de qualquer imagem foi o que travou a sala em 27/08 -- e a
    ordem das secoes nao tem nada a ver com isso.

    A peca .analogia carrega tres das cinco partes do AIDEN: a imagem (corpo),
    o detalhamento (.analogia-mapa, o de-para) e a extensao (.analogia-mais).
    O abstrato mora no .conceito e a negacao no .contraste, que ja existiam.
    """
    if not _e_aula(html):
        return []
    falhas = []
    for i, (_, corpo) in enumerate(blocos_por_classe(html, "conceito"), 1):
        if 'class="analogia"' not in corpo:
            falhas.append("{}: o conceito {} nao tem analogia. A imagem nao vem "
                          "depois do conceito: ela e como o conceito se diz"
                          .format(rel, i))
    for i, (_, corpo) in enumerate(blocos_por_classe(html, "analogia"), 1):
        if 'class="analogia-mapa"' not in corpo:
            falhas.append("{}: a analogia {} nao mapeia a imagem na coisa. Sem o "
                          "de-para ela e enfeite".format(rel, i))
    return falhas


def g40_a_aula_declara_onde_acaba(rel, html):
    """🔴 Toda aula tem o divisor, e ele fica entre o gabarito e as pegadinhas.

    A aula tem DOIS COMPRIMENTOS declarados: ate o divisor e obrigatorio, o
    resto e aprofundamento e o aluno escolhe. Peca HERDADA de "Claude para
    Lideres" (Adriano Couto) -- herdamos a forma, nada do texto.

    Medido na aula n1 do piloto IEL: 2.349 palavras ate o gabarito e 1.118
    depois, sendo 982 so na secao 08. Sem o divisor tudo chega como
    obrigatorio, e a aula de nivelamento vira 3.467 palavras de leitura.
    """
    if not _e_aula(html):
        return []
    if 'class="ate-aqui"' not in html:
        return ["{}: aula sem o divisor. Toda aula declara onde acaba o "
                "obrigatorio e comeca o aprofundamento".format(rel)]
    if 'class="ate-aqui-nota"' not in html:
        return ["{}: o divisor esta sem a nota que diz ao aluno o que fazer "
                "com o que vem depois".format(rel)]
    return []


# =========================================================================
# A LISTA · gate, defeito injetado, página alvo do defeito
# =========================================================================
GATES = [
    ("G1", "travessão", g1_travessao,
     lambda h: h.replace("<h1>", "<h1>defeito — injetado ", 1), None),
    ("G2", "vocabulário interno na tela", g2_vocabulario_interno,
     lambda h: h.replace("<h1>", "<h1>a onda 3 do handoff ", 1), "caso/index.html"),
    ("G3", "direção de cena", g3_direcao_de_cena,
     lambda h: h.replace("<h1>", "<h1>pergunte à sala e espere o silêncio ", 1),
     "caso/index.html"),
    ("G4", "duração de aula fora da capa", g4_minutagem_fora_da_capa,
     lambda h: h.replace("<h1>", "<h1>bloco de 15 min ", 1), "caso/index.html"),
    ("G5", "o prompt tem os quatro parágrafos", g5_prompt_tem_os_quatro_paragrafos,
     lambda h: re.sub(r'(<pre class="prompt-txt"[^>]*>[^<]*?)Na dúvida:',
                      r"\1Se precisar:", h, count=1),
     "caso/index.html"),
    ("G6", "classe sem CSS", g6_classe_sem_css,
     lambda h: h.replace('class="cartao', 'class="classe-inventada cartao', 1), None),
    ("G7", "os links resolvem", g7_links_resolvem,
     lambda h: h.replace('href="../exemplo/"', 'href="../pagina-que-nao-existe/"', 1),
     "caso/index.html"),
    ("G8", "as imagens existem no disco", g8_imagens_existem,
     lambda h: re.sub(r'(<img[^>]+src=")([^"]+)(")', r'\1../_img/nao-existe.svg\3', h, count=1),
     "componentes/index.html"),
    ("G9", "o gabarito nasce fechado", g9_gabarito_fechado,
     lambda h: re.sub(r'(<details[^>]*\bgabarito\b[^>]*?)>', r'\1 open>', h, count=1),
     "componentes/index.html"),
    ("G10", "o botão de copiar tem alvo", g10_botao_de_copiar_tem_alvo,
     lambda h: re.sub(r'data-alvo="[^"]+"', 'data-alvo="id-que-sumiu"', h, count=1),
     "caso/index.html"),
    ("G11", "a numeração é sequencial", g11_numeracao_sequencial,
     lambda h: h.replace('<div class="passo-n">03</div>',
                         '<div class="passo-n">07</div>', 1), "caso/index.html"),
    ("G12", "toda seção tem função e título", g12_secao_tem_funcao_e_titulo,
     lambda h: h.replace('class="secao-fn"', 'class="secao-fn-renomeada"', 1),
     "componentes/index.html"),
    ("G13", "tabela dentro do envelope", g13_tabela_no_envelope,
     lambda h: h.replace('<div class="tabela solta">', "<div>", 1),
     "componentes/index.html"),
    # O defeito tem de cair DENTRO de um <p>: a cola só alcança h1-4, <p> e <li>.
    ("G14", "a cola de quebra de linha está aplicada", g14_cola_aplicada,
     lambda h: re.sub(r'(<p\b[^>]*>(?:(?!</p>).)*?)&nbsp;', r'\1 ', h, count=1, flags=re.S),
     None),
    ("G15", "prosa sem quebra equilibrada", g15_prosa_sem_quebra_equilibrada,
     lambda h: h.replace("h1,h2,h3,h4{text-wrap:balance}",
                         "h1,h2,h3,h4{text-wrap:balance}\np,li{text-wrap:pretty}", 1),
     None),
    ("G16", "cor escrita só no marca.css", g16_cor_so_no_marca_css,
     lambda h: h.replace(".btn:hover{background:var(--accent-dark)",
                         ".btn:hover{background:#1A5670", 1), None),
    ("G17", "hex de 8 dígitos em SVG", g17_hex_de_oito_digitos,
     lambda h: h.replace("<body>", '<body><svg><rect fill="#26262322"></rect></svg>', 1),
     None),
    ("G18", "o criador de prompt está completo", g18_criador_completo,
     lambda h: h.replace('data-titulo="FORMATO"', 'data-titulo="SAÍDA"', 1),
     "componentes/index.html"),
    ("G19", "o canvas está completo", g19_canvas_completo,
     lambda h: h.replace('<div class="cv-n">4</div>', "", 1), "componentes/index.html"),
    ("G33", "a narrativa não tem o instrutor", g33_narrativa_sem_instrutor,
     # o .heroi-sub existe em toda página; <h2> não existe na página de caso,
     # e o injetor que não acha alvo deixa o gate cego sem nada ter mudado
     lambda h: h.replace('<p class="heroi-sub">',
                         '<p class="heroi-sub">Abro a conversa e mostro para vocês. ', 1),
     None),
    ("G34", "o .fr-host tem as frases marcadas", g34_fr_host_tem_as_frases_marcadas,
     # tira TODAS: o gate só reprova quando nenhuma frase está marcada, e
     # remover uma deixaria as outras seis passando
     lambda h: h.replace('<span class="fr">', ""),
     "componentes/index.html"),
    ("G32", "a página bate com o insumo", g32_a_pagina_bate_com_o_insumo,
     # defeito por PADRAO, nao pelo vocabulario do template: troca o nome da
     # primeira coluna citada, seja ela qual for no insumo deste curso.
     # a coluna em Maiuscula e a que o gate confere contra o insumo; trocar a
     # primeira minuscula que aparecer na prosa nao prova nada.
     lambda h: re.sub(r"(a coluna )([A-ZÀ-Ý]\w*)", r"\1Filial", h, count=1),
     None),
    ("G31", "o roteiro cita classe que existe", g31_o_roteiro_cita_classe_que_existe,
     lambda h: h.replace(".diagnostico{", ".classe-que-sumiu{", 1),
     "index.html"),
    ("G28", "a faixa cabe na régua", g28_faixa_cabe_na_regua,
     lambda h: h.replace('style="--de:6;--ate:6"', 'style="--de:6;--ate:9"', 1),
     "componentes/index.html"),
    ("G29", "o radar bate eixo com valor", g29_radar_bate_eixo_com_valor,
     lambda h: h.replace('data-valores="35,80,55,30,45"', 'data-valores="35,80,55,30"', 1),
     "componentes/index.html"),
    ("G30", "o mapa de área soma cem", g30_mapa_de_area_soma_cem,
     lambda h: h.replace('<div class="ma" style="--w:25">',
                         '<div class="ma" style="--w:35">', 1),
     "componentes/index.html"),
    ("G27", "grade de N colunas pode encolher", g27_grade_de_n_colunas_pode_encolher,
     lambda h: h.replace("repeat(var(--n),minmax(0,1fr))", "repeat(var(--n),1fr)", 1),
     None),
    ("G25", "o cem bate com a legenda", g25_o_cem_bate_com_a_legenda,
     lambda h: h.replace('<i class="cem-p aceso"></i>', '<i class="cem-p"></i>', 1),
     "componentes/index.html"),
    ("G26", "as colunas têm três listas iguais", g26_colunas_com_tres_listas_iguais,
     lambda h: h.replace('<span class="apagado">Parado</span>', '', 1),
     "componentes/index.html"),
    ("G23", "pergunta de grupo tem destrave", g23_pergunta_de_grupo_tem_destrave,
     lambda h: re.sub(r'<div class="destrave">.*?</div>\s*</div>\s*</div>', '</div>',
                      h, count=1, flags=re.S),
     "componentes/index.html"),
    ("G24", "os dois loops começam iguais", g24_os_dois_loops_comecam_iguais,
     lambda h: h.replace('<span class="loop-no">Alguém decide</span>',
                         '<span class="loop-no">Alguem decide</span>', 1),
     "componentes/index.html"),
    ("G22", "a matriz tem os quatro quadrantes", g22_matriz_com_os_quatro_quadrantes,
     lambda h: h.replace('<div class="quad sobrar">', '<div class="quad">', 1),
     "componentes/index.html"),
    ("G21", "o script não está quebrado", g21_o_script_nao_esta_quebrado,
     lambda h: h.replace("var t = document.createElement('textarea');",
                         "var t = document.createElement('textarea);", 1), None),
    ("G20", "o exemplo pronto imprime sozinho", g20_exemplo_imprime_sozinho,
     lambda h: h.replace("@media print", "@media screen and (min-width:99999px)", 1),
     "exemplo/index.html"),
    ("G35", "o breakout nao sobrevive a shorthand de margem",
     g35_breakout_nao_sobrevive_a_shorthand,
     lambda h: h.replace(".aulas-lista{display:flex",
                         ".aulas-lista{margin:22px 0;display:flex", 1),
     None),
    ("G36", "a mesa fecha 100", g36_a_mesa_fecha_cem,
     # defeito por PADRAO: desloca a largura da primeira faixa, qualquer que
     # seja o valor dela nesta pagina, e a soma deixa de fechar 100.
     lambda h: re.sub(r'(class="mesa-faixa[^"]*"[^>]*style="width:)([\d.]+)(%)',
                      lambda m: "{}{:g}{}".format(m.group(1),
                                                  float(m.group(2)) + 7, m.group(3)),
                      h, count=1),
     None),
    ("G37", "as divs fecham", g37_as_divs_fecham,
     lambda h: h.replace("</div>", "", 1), None),
    ("G38", "todo bloco mora em envelope", g38_todo_bloco_mora_em_envelope,
     lambda h: re.sub(r"(<main\b[^>]*>)",
                      r'\1<div class="cartao">defeito injetado</div>', h, count=1),
     None),
    ("G39", "uma aula, um conceito", g39_uma_aula_um_conceito,
     lambda h: h.replace("</main>",
                         '<div class="conceito">defeito injetado</div></main>', 1),
     None),
    ("G40", "a aula declara onde acaba", g40_a_aula_declara_onde_acaba,
     lambda h: h.replace('class="ate-aqui"', 'class="ate-aqui-sumiu"', 1), None),
    ("G41", "o conceito vem com imagem", g41_o_conceito_vem_com_imagem,
     lambda h: h.replace('class="analogia"', 'class="analogia-sumiu"', 1), None),
]


def main():
    docs = paginas()
    if not docs:
        sys.exit("nenhuma página encontrada em " + RAIZ)

    print("Gates do padrão · {} páginas\n".format(len(docs)))
    print("  " + " · ".join(rel for rel, _ in docs) + "\n")

    por_rel = dict(docs)
    total_falhas = total_checagens = 0
    cegos = []

    for gid, nome, fn, defeito, alvo_defeito in GATES:
        falhas = []
        for rel, html in docs:
            total_checagens += 1
            for f in fn(rel, html):
                falhas.append("{}: {}".format(rel, f))

        # 🔴 A CALIBRACAO ACHA O ALVO SOZINHA.
        # Ate 28/08 todo alvo era NOME de pagina do template ("modulo/index.html").
        # Nenhum curso real usa esses nomes: no piloto IEL as paginas se chamavam
        # "nivelamento" e "n1-dois-modos", e tres gates nasceram cegos de uma vez
        # -- a suite reprovou inteira sem um achado real, e a sessao teve que
        # reapontar os alvos a mao, conserto que morre na proxima sincronizacao.
        # Agora o alvo declarado e so uma DICA: se ele nao existe, ou se o defeito
        # nao prova nada nele, a calibracao varre as demais paginas e fica na
        # primeira em que o defeito injetado realmente faz o gate acusar.
        # Gate novo pode nascer com alvo None: ele se acha.
        candidatos = ([alvo_defeito] if alvo_defeito in por_rel else []) \
            + [r for r, _ in docs if r != alvo_defeito]
        alvo, acusou = alvo_defeito or candidatos[0], False
        for cand in candidatos:
            limpo = por_rel[cand]
            sujo = defeito(limpo)
            if sujo == limpo:
                continue
            if len(fn(cand, sujo)) > len(fn(cand, limpo)):
                alvo, acusou = cand, True
                break
        if not acusou:
            cegos.append("{} · nenhuma das {} páginas prova este gate: ou o defeito "
                         "injetado não muda nada, ou o gate não o acusa"
                         .format(gid, len(candidatos)))

        marca = "FALHA " if falhas else ("CEGO  " if not acusou else "ok    ")
        print("{} {:<4} {:<40} {} achado(s) · calibrado: {}".format(
            marca, gid, nome, len(falhas), "sim" if acusou else "NÃO"))
        for f in falhas[:12]:
            print("        " + f)
        if len(falhas) > 12:
            print("        ... e mais {}".format(len(falhas) - 12))
        total_falhas += len(falhas)

    print("\n{} checagens · {} achados".format(total_checagens, total_falhas))

    if cegos:
        print("\nGATES QUE NÃO SE PROVARAM (não valem como verificação):")
        for g in cegos:
            print("  " + g)

    if total_falhas or cegos:
        print("\nRESULTADO: FALHA")
        sys.exit(1)
    print("\nRESULTADO: passou")


if __name__ == "__main__":
    main()
